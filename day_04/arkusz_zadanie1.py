# Biblioteka do stylizacji/formatowanie textu
import openpyxl

workbook = openpyxl.load_workbook('sales.xlsx')
worksheet = workbook.active

print(workbook)

print(worksheet)  # <Worksheet "Arkusz1">

lista = []
for w in worksheet:
    print(w)

for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        lista.append(col[i].value)

print("------ LISTA --------")
print(lista)

for i in range(0, len(lista)):
    print(lista[i:i + 3])
